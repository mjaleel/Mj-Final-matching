import streamlit as st
import pandas as pd
import re
from rapidfuzz import fuzz
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from io import BytesIO
import random

# ============ دوال المساعدة ============

def generate_color():
    colors = [
        "FFB3BA", "BAFFC9", "BAE1FF", "FFFFBA", "FFDFBA", "E0BBE4",
        "957DAD", "D4A5A5", "A8E6CF", "DCEDC1", "FFD3B6", "FFAAA5",
        "FF8B94", "A8D8EA", "AA96DA", "FCBAD3", "C9CBA3", "FFE66D",
        "F38181", "95E1D3", "EAFFD0", "FCE38A", "F54748", "7FE7CC"
    ]
    return random.choice(colors)

def normalize_name(name):
    if pd.isnull(name):
        return ""
    name = str(name).strip()
    name = name.replace("ه", "ة").replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    name = name.replace("ى", "ي").replace("ئ", "ي")
    name = re.sub(r'(عبد)([^\s])', r'\1 \2', name)
    return " ".join(name.split()).lower()

def get_first_three_words(name):
    if pd.isnull(name) or name == "":
        return ""
    words = str(name).split()
    return " ".join(words[:3]) if len(words) >= 3 else " ".join(words)

def is_first_three_words_match(name1, name2):
    words1 = name1.split()
    words2 = name2.split()
    length = min(len(words1), len(words2), 3)
    if length == 0:
        return False
    return all(words1[i] == words2[i] for i in range(length))

# ============ دوال المطابقة ============

def match_names(names_df, database_df, name_column_file, name_column_db, selected_columns=None):
    names_df = names_df.copy()
    database_df = database_df.copy()
    
    names_df["normalized_name"] = names_df[name_column_file].apply(normalize_name)
    database_df["normalized_name"] = database_df[name_column_db].apply(normalize_name)
    database_df = database_df.drop_duplicates(subset=["normalized_name"])
    
    columns_to_get = [name_column_db]
    if selected_columns:
        columns_to_get.extend([col for col in selected_columns if col != name_column_db])
    
    database_map = database_df.set_index("normalized_name")[columns_to_get].to_dict(orient="index")
    matched_results = []

    for original_name, normalized_name in zip(names_df[name_column_file], names_df["normalized_name"]):
        best_match = None
        best_score = 0
        
        for db_name in database_map.keys():
            score = fuzz.ratio(normalized_name, db_name)
            if score > best_score:
                best_score = score
                best_match = db_name

        match_data = None
        if best_score >= 85 and best_match:
            if is_first_three_words_match(normalized_name, best_match) or best_match.startswith(normalized_name):
                match_data = database_map[best_match]
        
        if not match_data:
            for db_name in database_map.keys():
                if db_name.startswith(normalized_name) or normalized_name.startswith(db_name):
                    match_data = database_map[db_name]
                    best_match = db_name
                    best_score = fuzz.ratio(normalized_name, best_match)
                    break

        if match_data:
            result = {
                "الاسم الأصلي": original_name,
                "الاسم المطابق": match_data[name_column_db],
                "نسبة التطابق": f"{round(best_score)}%",
                "ملاحظة": "✅ تطابق"
            }
            if selected_columns:
                for col in selected_columns:
                    if col in match_data:
                        result[col] = match_data[col]
            matched_results.append(result)
        else:
            result = {
                "الاسم الأصلي": original_name,
                "الاسم المطابق": "",
                "نسبة التطابق": "",
                "ملاحظة": "❌ لم يتم العثور على تطابق"
            }
            if selected_columns:
                for col in selected_columns:
                    result[col] = ""
            matched_results.append(result)

    return pd.DataFrame(matched_results)


def match_three_word_names(
    names_df, database_df,
    name_column_file, name_column_db,
    selected_columns=None,
    school_column_file=None,   # عمود المدرسة في ملف الأسماء
    school_column_db=None      # عمود المدرسة في قاعدة البيانات
):
    """
    مطابقة ثلاثية مع دعم المطابقة المزدوجة (اسم + مدرسة).
    - إذا تم تحديد عمودَي المدرسة: المطابقة تشترط تطابق الاسم الثلاثي والمدرسة معاً.
    - إذا لم يتم تحديدهما: تعمل كالسابق بالاسم الثلاثي فقط.
    """
    use_school = bool(school_column_file and school_column_db)

    names_df = names_df.copy()
    database_df = database_df.copy()

    # ---- تطبيع الأسماء ----
    names_df["normalized_name"] = names_df[name_column_file].apply(normalize_name)
    names_df["three_word_name"] = names_df["normalized_name"].apply(get_first_three_words)

    database_df["normalized_name"] = database_df[name_column_db].apply(normalize_name)
    database_df["three_word_name"] = database_df["normalized_name"].apply(get_first_three_words)

    # ---- تطبيع المدرسة إن وُجدت ----
    if use_school:
        names_df["normalized_school"] = names_df[school_column_file].apply(normalize_name)
        database_df["normalized_school"] = database_df[school_column_db].apply(normalize_name)

    # ---- ألوان الأسماء الثلاثية المشتركة ----
    file_three_words = set(names_df["three_word_name"].dropna().unique())
    db_three_words = set(database_df["three_word_name"].dropna().unique())
    common_three_words = file_three_words.intersection(db_three_words)

    color_map = {}
    for name in common_three_words:
        if name and name not in color_map:
            color_map[name] = generate_color()

    matched_results = []

    for idx, row in names_df.iterrows():
        original_name = row[name_column_file]
        three_word    = row["three_word_name"]
        normalized    = row["normalized_name"]
        school_norm   = row["normalized_school"] if use_school else None

        # ---- تصفية قاعدة البيانات بالاسم الثلاثي أولاً ----
        name_matches = database_df[database_df["three_word_name"] == three_word]

        if use_school and len(name_matches) > 0:
            # ---- تصفية ثانية بالمدرسة (نفس منطق fuzzy) ----
            school_matches = []
            for _, db_row in name_matches.iterrows():
                db_school = db_row["normalized_school"]
                score_school = fuzz.ratio(school_norm, db_school)
                # نفس منطق الاسم: 85% أو startswith
                if (
                    score_school >= 85
                    or db_school.startswith(school_norm)
                    or school_norm.startswith(db_school)
                ):
                    school_matches.append(db_row)

            final_matches = pd.DataFrame(school_matches) if school_matches else pd.DataFrame()
            match_type = "✅ تطابق اسم + مدرسة"
            no_match_note = self_note(three_word, name_matches)
        else:
            final_matches = name_matches
            match_type = "✅ تطابق ثلاثي"
            no_match_note = "❌ لا يوجد تطابق"

        if len(final_matches) > 0:
            for _, match_row in final_matches.iterrows():
                score = fuzz.ratio(normalized, match_row["normalized_name"])
                result = {
                    "الاسم الأصلي": original_name,
                    "الاسم الثلاثي": three_word,
                    "الاسم المطابق": match_row[name_column_db],
                    "نسبة التطابق": f"{round(score)}%",
                    "ملاحظة": match_type,
                    "_color": color_map.get(three_word, "FFFFFF")
                }
                if use_school:
                    result["المدرسة (الملف)"]    = row[school_column_file]
                    result["المدرسة (قاعدة البيانات)"] = match_row[school_column_db]
                if selected_columns:
                    for col in selected_columns:
                        if col in match_row.index and col not in (name_column_db, school_column_db):
                            result[col] = match_row[col]
                matched_results.append(result)
        else:
            # لا تطابق — وضّح السبب
            if use_school and len(name_matches) > 0:
                note = "⚠️ اسم متطابق لكن المدرسة مختلفة"
            else:
                note = "❌ لا يوجد تطابق"

            result = {
                "الاسم الأصلي": original_name,
                "الاسم الثلاثي": three_word,
                "الاسم المطابق": "",
                "نسبة التطابق": "",
                "ملاحظة": note,
                "_color": "FFCCCC" if "⚠️" in note else "FFFFFF"
            }
            if use_school:
                result["المدرسة (الملف)"]    = row[school_column_file]
                result["المدرسة (قاعدة البيانات)"] = ""
            if selected_columns:
                for col in selected_columns:
                    result[col] = ""
            matched_results.append(result)

    return pd.DataFrame(matched_results), color_map


def self_note(three_word, name_matches):
    """دالة مساعدة داخلية — غير مستخدمة خارجياً"""
    return "❌ لا يوجد تطابق"


def find_similar_groups(names_df, name_column):
    names_df = names_df.copy()
    names_df["normalized_name"] = names_df[name_column].apply(normalize_name)
    names_df["three_word_name"] = names_df["normalized_name"].apply(get_first_three_words)
    
    three_word_counts = names_df["three_word_name"].value_counts()
    duplicates = three_word_counts[three_word_counts > 1].index.tolist()
    
    color_map = {}
    for name in duplicates:
        if name:
            color_map[name] = generate_color()
    
    names_df["متكرر"] = names_df["three_word_name"].apply(lambda x: "✅ متكرر" if x in duplicates else "")
    names_df["_color"] = names_df["three_word_name"].apply(lambda x: color_map.get(x, "FFFFFF"))
    
    return names_df, color_map

# ============ دوال التصدير ============

def to_excel_basic(df):
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active
        red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")

        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            col_letter = column[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        note_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if "ملاحظة" in str(cell.value):
                note_col_idx = idx
                break
        
        if note_col_idx:
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                if row[note_col_idx-1].value and "❌" in str(row[note_col_idx-1].value):
                    for cell in row:
                        cell.fill = red_fill

    output.seek(0)
    return output

def to_excel_with_colors(df, color_map):
    output = BytesIO()
    df_export = df.drop(columns=['_color'], errors='ignore')
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False)
        wb = writer.book
        ws = wb.active
        red_fill    = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

        for column in ws.columns:
            max_length = max(len(str(cell.value)) if cell.value else 0 for cell in column)
            col_letter = column[0].column_letter
            ws.column_dimensions[col_letter].width = max_length + 2

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
            if row_idx < len(df):
                note  = df.iloc[row_idx].get("ملاحظة", "")
                color = df.iloc[row_idx].get("_color", "FFFFFF")
                
                if "❌" in str(note):
                    for cell in row:
                        cell.fill = red_fill
                elif "⚠️" in str(note):
                    for cell in row:
                        cell.fill = yellow_fill
                elif color and color != "FFFFFF":
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    for cell in row:
                        cell.fill = fill

    output.seek(0)
    return output

def update_file_names(original_df, results_df, name_column, update_column="الاسم المطابق"):
    updated_df = original_df.copy()
    for idx, row in results_df.iterrows():
        if "✅" in str(row.get("ملاحظة", "")) and row.get(update_column):
            original_normalized = normalize_name(row["الاسم الأصلي"])
            for i, orig_row in updated_df.iterrows():
                if normalize_name(orig_row[name_column]) == original_normalized:
                    updated_df.at[i, name_column] = row[update_column]
                    break
    return updated_df

# ============ واجهة المستخدم ============

st.set_page_config(page_title="تطبيق المطابقة المتقدم", layout="wide")
st.markdown("### 👨‍💻 برمجة: محمد عبدالجليل")
st.title("🔐 نظام مطابقة الأسماء المتقدم")

password = st.text_input("أدخل كلمة المرور:", type="password")

if password == "mjaleel":
    st.success("✅ تم التحقق من كلمة المرور.")
    
    tab1, tab2, tab3 = st.tabs([
        "📋 مطابقة الأسماء", 
        "🏢 مطابقة الأقسام", 
        "🔄 مطابقة الأسماء الثلاثية"
    ])

    # ============ تبويب 1 ============
    with tab1:
        st.subheader("📋 مطابقة الأسماء العامة")
        col1, col2 = st.columns(2)
        with col1:
            file1 = st.file_uploader("📄 ملف الأسماء", type="xlsx", key="file1_name")
        with col2:
            file2 = st.file_uploader("📊 ملف قاعدة البيانات", type="xlsx", key="file2_name")

        if file1 and file2:
            names_df = pd.read_excel(file1)
            db_df    = pd.read_excel(file2)
            
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                name_col_file = st.selectbox("🔹 عمود الاسم في ملف الأسماء:", names_df.columns.tolist(), key="name_col_file_tab1")
            with col2:
                name_col_db = st.selectbox("🔹 عمود الاسم في قاعدة البيانات:", db_df.columns.tolist(), key="name_col_db_tab1")
            
            available_columns = [col for col in db_df.columns.tolist() if col != name_col_db]
            selected_columns  = st.multiselect("📌 اختر الأعمدة المراد جلبها:", available_columns, key="cols_tab1")
            
            if st.button("🚀 بدء المطابقة", key="match_names_btn"):
                with st.spinner("جاري المطابقة..."):
                    results = match_names(names_df, db_df, name_col_file, name_col_db, selected_columns)
                st.success("✅ تم المطابقة بنجاح!")
                total   = len(results)
                matched = len(results[results["ملاحظة"].str.contains("✅", na=False)])
                col1, col2, col3 = st.columns(3)
                col1.metric("📊 الإجمالي", total)
                col2.metric("✅ متطابق", matched)
                col3.metric("❌ غير متطابق", total - matched)
                st.dataframe(results, use_container_width=True)
                st.download_button("⬇️ تحميل النتائج", to_excel_basic(results), file_name="نتائج_المطابقة.xlsx")

    # ============ تبويب 2 ============
    with tab2:
        st.subheader("🏢 مطابقة الأقسام والإدارات")
        col1, col2 = st.columns(2)
        with col1:
            file3 = st.file_uploader("📄 ملف الأسماء", type="xlsx", key="file1_dept")
        with col2:
            file4 = st.file_uploader("📊 ملف قاعدة البيانات", type="xlsx", key="file2_dept")

        if file3 and file4:
            names_df = pd.read_excel(file3)
            db_df    = pd.read_excel(file4)
            
            st.markdown("---")
            col1, col2 = st.columns(2)
            with col1:
                name_col_file = st.selectbox("🔹 عمود الاسم في ملف الأسماء:", names_df.columns.tolist(), key="name_col_file_tab2")
            with col2:
                name_col_db = st.selectbox("🔹 عمود الاسم في قاعدة البيانات:", db_df.columns.tolist(), key="name_col_db_tab2")
            
            available_columns = [col for col in db_df.columns.tolist() if col != name_col_db]
            selected_columns  = st.multiselect("📌 اختر الأعمدة المراد جلبها:", available_columns, key="cols_tab2")
            
            if st.button("🚀 بدء المطابقة", key="match_dept_btn"):
                with st.spinner("جاري المطابقة..."):
                    results = match_names(names_df, db_df, name_col_file, name_col_db, selected_columns)
                st.success("✅ تم المطابقة بنجاح!")
                total   = len(results)
                matched = len(results[results["ملاحظة"].str.contains("✅", na=False)])
                col1, col2, col3 = st.columns(3)
                col1.metric("📊 الإجمالي", total)
                col2.metric("✅ متطابق", matched)
                col3.metric("❌ غير متطابق", total - matched)
                st.dataframe(results, use_container_width=True)
                st.download_button("⬇️ تحميل النتائج", to_excel_basic(results), file_name="نتائج_الأقسام.xlsx")

    # ============ تبويب 3 — المطابقة الثلاثية مع المدرسة ============
    with tab3:
        st.subheader("🔄 مطابقة الأسماء الثلاثية")
        st.info("""
        💡 **هذا التبويب يدعم:**
        - مطابقة أول 3 كلمات من الاسم
        - مطابقة مزدوجة: **اسم + مدرسة** (اختياري)
        - تلوين المجموعات المتشابهة
        - جلب أي أعمدة إضافية ديناميكياً
        """)

        col1, col2 = st.columns(2)
        with col1:
            file5 = st.file_uploader("📄 ملف الأسماء", type="xlsx", key="file1_three")
        with col2:
            file6 = st.file_uploader("📊 ملف قاعدة البيانات", type="xlsx", key="file2_three")

        if file5 and file6:
            names_df = pd.read_excel(file5)
            db_df    = pd.read_excel(file6)

            st.markdown("---")
            st.subheader("⚙️ إعدادات الأعمدة")

            # ---- اختيار عمود الاسم ----
            col1, col2 = st.columns(2)
            with col1:
                name_col_file = st.selectbox("🔹 عمود الاسم في ملف الأسماء:", names_df.columns.tolist(), key="name_col_file_tab3")
            with col2:
                name_col_db = st.selectbox("🔹 عمود الاسم في قاعدة البيانات:", db_df.columns.tolist(), key="name_col_db_tab3")

            # ---- مطابقة المدرسة (اختياري) ----
            st.markdown("---")
            use_school = st.checkbox("🏫 تفعيل مطابقة المدرسة (اسم + مدرسة معاً)", key="use_school_check")

            school_col_file = None
            school_col_db   = None

            if use_school:
                col1, col2 = st.columns(2)
                with col1:
                    # الأعمدة المتبقية في ملف الأسماء (غير عمود الاسم)
                    remaining_file_cols = [c for c in names_df.columns.tolist() if c != name_col_file]
                    school_col_file = st.selectbox("🔹 عمود المدرسة في ملف الأسماء:", remaining_file_cols, key="school_col_file")
                with col2:
                    remaining_db_cols = [c for c in db_df.columns.tolist() if c != name_col_db]
                    school_col_db = st.selectbox("🔹 عمود المدرسة في قاعدة البيانات:", remaining_db_cols, key="school_col_db")

            # ---- اختيار الأعمدة الإضافية ديناميكياً ----
            st.markdown("---")
            # استبعاد عمود الاسم وعمود المدرسة من الاختيار
            excluded = {name_col_db}
            if school_col_db:
                excluded.add(school_col_db)
            available_columns = [c for c in db_df.columns.tolist() if c not in excluded]

            selected_columns = st.multiselect(
                "📌 اختر الأعمدة المراد جلبها من قاعدة البيانات:",
                available_columns,
                default=available_columns[:3] if len(available_columns) >= 3 else available_columns,
                key="cols_tab3"
            )

            st.markdown("---")

            # ---- خيارات إضافية ----
            col1, col2 = st.columns(2)
            with col1:
                update_names_option = st.checkbox("✏️ تحديث الأسماء في الملف الأصلي", key="update_names_check")
            with col2:
                show_similar_only = st.checkbox("👁️ عرض المتطابقات فقط", key="show_similar_only")

            st.markdown("---")

            if st.button("🚀 بدء المطابقة الثلاثية", key="match_three_btn", type="primary"):
                with st.spinner("جاري المطابقة..."):
                    results, color_map = match_three_word_names(
                        names_df, db_df,
                        name_col_file, name_col_db,
                        selected_columns,
                        school_column_file=school_col_file,
                        school_column_db=school_col_db
                    )

                st.success("✅ تم المطابقة بنجاح!")

                # ---- إحصائيات ----
                total    = len(results)
                matched  = len(results[results["ملاحظة"].str.contains("✅", na=False)])
                warned   = len(results[results["ملاحظة"].str.contains("⚠️", na=False)])
                no_match = len(results[results["ملاحظة"].str.contains("❌", na=False)])
                groups   = len(color_map)

                if use_school:
                    col1, col2, col3, col4, col5 = st.columns(5)
                    col1.metric("📊 الإجمالي", total)
                    col2.metric("✅ اسم + مدرسة", matched)
                    col3.metric("⚠️ اسم فقط", warned)
                    col4.metric("❌ لا تطابق", no_match)
                    col5.metric("🎨 مجموعات", groups)
                else:
                    col1, col2, col3, col4 = st.columns(4)
                    col1.metric("📊 الإجمالي", total)
                    col2.metric("✅ متطابق", matched)
                    col3.metric("❌ غير متطابق", total - matched)
                    col4.metric("🎨 مجموعات", groups)

                # ---- عرض النتائج ----
                display_results = results.drop(columns=['_color'], errors='ignore')
                if show_similar_only:
                    display_results = display_results[display_results["ملاحظة"].str.contains("✅", na=False)]

                st.dataframe(display_results, use_container_width=True)

                st.download_button(
                    "⬇️ تحميل نتائج المطابقة (ملون)",
                    to_excel_with_colors(results, color_map),
                    file_name="نتائج_المطابقة_الثلاثية.xlsx",
                    key="download_three_results"
                )

                st.markdown("---")

                # ---- تحديث الأسماء ----
                if update_names_option:
                    st.subheader("📝 تحديث الملف الأصلي")
                    with st.spinner("جاري التحديث..."):
                        updated_df = update_file_names(names_df, results, name_col_file, "الاسم المطابق")

                    col1, col2 = st.columns(2)
                    with col1:
                        st.markdown("**📄 الملف المحدث:**")
                        st.dataframe(updated_df, use_container_width=True)
                        output_updated = BytesIO()
                        with pd.ExcelWriter(output_updated, engine='openpyxl') as writer:
                            updated_df.to_excel(writer, index=False)
                        output_updated.seek(0)
                        st.download_button("⬇️ تحميل الملف المحدث", output_updated, file_name="الملف_المحدث.xlsx", key="download_updated_file")
                    with col2:
                        st.markdown("**📊 ملخص التحديثات:**")
                        changes = results[results["ملاحظة"].str.contains("✅", na=False)]
                        st.dataframe(changes[["الاسم الأصلي", "الاسم المطابق"]].drop_duplicates(), use_container_width=True)

                # ---- مجموعات الأسماء المتشابهة ----
                st.markdown("---")
                st.subheader("🎨 مجموعات الأسماء المتشابهة")
                if color_map:
                    for three_word, color in color_map.items():
                        group_data = results[results["الاسم الثلاثي"] == three_word]
                        if len(group_data) > 0:
                            with st.expander(f"👥 {three_word} ({len(group_data)} أسماء)"):
                                st.dataframe(
                                    group_data[["الاسم الأصلي", "الاسم المطابق"]].drop_duplicates(),
                                    use_container_width=True
                                )
                else:
                    st.info("لا توجد أسماء متشابهة")

elif password:
    st.error("❌ كلمة المرور غير صحيحة.")
